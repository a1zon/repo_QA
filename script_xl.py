from openpyxl import load_workbook

workbook = load_workbook(filename="tmp/file_example_XLSX_1000.xlsx")  # загрузка 
sheet = workbook.active  # активный лист
print(sheet.cell(row =2, column=2).value)  # чтение ячейки
for x in sheet.columns: 
    for cell in x:
        print(cell.value)